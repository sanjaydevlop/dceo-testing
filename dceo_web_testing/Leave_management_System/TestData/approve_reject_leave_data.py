import openpyxl,os,sys
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, os.pardir))
grandparent_dir = os.path.abspath(os.path.join(parent_dir, os.pardir))
# Add the grandparent directory to sys.path
sys.path.append(grandparent_dir)
class approve_reject_leave_data:
    def get_approve_reject_leave_data(test_sheet_name):
        workbook=openpyxl.load_workbook(f'{grandparent_dir}/Leave_management_System/TestData/Leader_Leave_Requests_Leave_Management_System.xlsx')
        sheet = workbook[test_sheet_name]
        max_cols = sheet.max_column
        dataPoints=[]
        for row in sheet.iter_rows(min_row=2,max_row=500,values_only=True):
            rowData=[]
            for i in range(1,max_cols):
                if(row[i]==None):
                    break
                else:
                    rowData.append(row[i])
            if(len(rowData)>0):
                dataPoints.append(tuple(rowData))
        return dataPoints
print(approve_reject_leave_data.get_approve_reject_leave_data("Sheet6"))


                



