# import openpyxl,os,sys
# current_dir = os.path.dirname(os.path.abspath(__file__))
# parent_dir = os.path.abspath(os.path.join(current_dir, os.pardir))
# grandparent_dir = os.path.abspath(os.path.join(parent_dir, os.pardir))
# # Add the grandparent directory to sys.path
# sys.path.append(grandparent_dir)
# class customer_ref_data:
#     def get_customer_ref_data(test_sheet_name):
#         workbook=openpyxl.load_workbook(f'{grandparent_dir}/Customer_Referral/TestData/customer_referral.xlsx')
#         sheet = workbook[test_sheet_name]
#         max_cols = sheet.max_column
#         dataPoints=[]
        
#         for row in sheet.iter_rows(min_row=2,max_row=500,values_only=True):
#             if(row[0]==None):
#                 break
#             rowData=[]
#             for i in range(1,max_cols):
#                 rowData.append(row[i])
#             dataPoints.append(tuple(rowData))
#         print(len(dataPoints))
#         return dataPoints
        
# print(customer_ref_data.get_customer_ref_data("Sheet3"))

import openpyxl,os,sys
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, os.pardir))
grandparent_dir = os.path.abspath(os.path.join(parent_dir, os.pardir))
sys.path.append(grandparent_dir)
class my_profile_data:
    def get_my_profile_data(testcase):
        workbook=openpyxl.load_workbook(f'{grandparent_dir}/My_Profile_Production/TestData/my_profile_production.xlsx')
        sheet = workbook["Sheet1"]
        dataPoints=[]
        filtered_rows = [row for row in sheet.iter_rows(min_row=2) if row[0].value == testcase]
        for row in filtered_rows:
            dataPoints.append(tuple([cell.value for cell in row[1:]]))
        return dataPoints
        
print(my_profile_data.get_my_profile_data("test_update_personal_details"))